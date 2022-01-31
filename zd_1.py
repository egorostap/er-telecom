# необходимо установить pandas, openpyxl
import os
import sqlite3
import openpyxl
import pandas as pd

# эксель таблицы загружены в бд sqlite, данная база данных также находится в репозитории проекта
# готовые запросы по домашнему заданию находятся в переменных: "zd_1_1 - zd_1_3"
def export_to_sqlite(xl='data_.xlsx', lst_xl=''):
    '''Экспорт данных из xlsx в sqlite'''

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 's.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    # cursor.execute(f'CREATE TABLE IF NOT EXISTS {lst_xl} (cl_id int, calc_period int, channel_id int , campaign_theme text, group_pos_1 int)')
    # cursor.execute(f'CREATE TABLE IF NOT EXISTS {lst_xl} (cl_id int, calc_period int, revenue_summ int , cl_activ int)')
    cursor.execute(f'CREATE TABLE IF NOT EXISTS {lst_xl} (cl_id int, product_name text, product_probability int)')


    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook(xl, data_only=True)
    sheet = file_to_read[lst_xl]

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 4):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)
        print(data)

    # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute(f"INSERT INTO {lst_xl} VALUES (?, ?, ?);", (data[0], data[1], data[2]))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base(table):
    '''Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 's.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute(f"DELETE FROM {table}")
    connect.commit()
    connect.close()


def input_base(sql_request):
    '''SQL запрос'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 's.sqlite3'
    # соединение с базой
    connect = sqlite3.connect(prj_dir + '/' + base_name)

    # Запрос в базу, сохранение и закрытие соединения
    df = pd.read_sql(sql_request, connect)
    # запись в отдельную таблицу бд без зависимостей
    # df_anythink.to_sql(“name_table”, connect)
    connect.commit()
    connect.close()
    print(df)
    return df

# переменные с наименованиями вкладок ексель
s1, s2, s3 = 's_1', 's_2', 's_3'
p1, p2, p3, p4 = 'p_1', 'p_2', 'p_3', 'p_4'

# переменные с решениями
zd_1_1 = '''
select distinct s_1.cl_id,
(select sum(revenue_summ)/count(cl_id) as arpu
from s_1 join s_2 using(cl_id)
where s_1.calc_period IN (202106, 202107, 202108)
and campaign_theme = "Продажа" 
and group_pos_1 = 1
and s_2.calc_period IN (202106, 202107, 202108)
and cl_activ = 1) as arpu,
(case when (rev_last.revenue_summ - rev_first.revenue_summ) > 0 then "True" else "False" end) as revenue_up

from s_1 
join s_2 using(cl_id)
join (select s_1.cl_id, min(s_2.calc_period), revenue_summ
from s_1 join s_2 using(cl_id)
where s_1.calc_period IN (202106, 202107, 202108)
and campaign_theme = "Продажа" 
and group_pos_1 = 1
and s_2.calc_period IN (202106, 202107, 202108)
and cl_activ = 1
group by s_1.cl_id) rev_first using(cl_id)
join (select s_1.cl_id, max(s_2.calc_period), revenue_summ
from s_1 join s_2 using(cl_id)
where s_1.calc_period IN (202106, 202107, 202108)
and campaign_theme = "Продажа" 
and group_pos_1 = 1
and s_2.calc_period IN (202106, 202107, 202108)
and cl_activ = 1
group by s_1.cl_id) rev_last using(cl_id)

where  s_1.calc_period IN (202106, 202107, 202108)
and campaign_theme = "Продажа" 
and group_pos_1 = 1
and s_2.calc_period IN (202106, 202107, 202108)
and cl_activ = 1
'''

zd_1_3 = '''
WITH A AS
(
select cl_id, product_name, product_probability,
ROW_NUMBER() over (PARTITION by cl_id order by product_probability desc) as rownum
from s_3
)
SELECT cl_id, product_name, sum(product_probability)/count(product_probability) as average_top_3_prod
FROM A
WHERE rownum <= 3
group by cl_id
'''

zd_1_2 = '''
WITH t_0 AS
(
SELECT campaign_theme, count(channel_id) as ch_0 from s_1 where channel_id = 0 group by campaign_theme
),
t_1 AS
(
SELECT campaign_theme, count(channel_id) as ch_1 from s_1 where channel_id = 1 group by campaign_theme
),
t_2 AS 
(
SELECT campaign_theme, count(channel_id) as ch_2 from s_1 where channel_id = 2 group by campaign_theme
),
t_3 AS 
(
SELECT campaign_theme, count(channel_id) as ch_3 from s_1 where channel_id = 3 group by campaign_theme
),
t_4 AS 
(
SELECT campaign_theme, count(channel_id) as ch_4 from s_1 where channel_id = 4 group by campaign_theme
),
t_5 AS 
(
SELECT campaign_theme, count(channel_id) as ch_5 from s_1 where channel_id = 5 group by campaign_theme
),
t_6 AS 
(
SELECT campaign_theme, count(channel_id) as ch_6 from s_1 where channel_id = 6 group by campaign_theme
)
SELECT t_0.campaign_theme, ch_0, ch_1, ch_2, ch_3, ch_4, ch_5, ch_6
FROM t_0
left join t_1 using (campaign_theme)
left join t_2 using (campaign_theme)
left join t_3 using (campaign_theme)
left join t_4 using (campaign_theme)
left join t_5 using (campaign_theme)
left join t_6 using (campaign_theme)

'''

# Запуск функции
if __name__ == '__main__':
    # pass
    input_base(sql_request=zd_1_3)
    # export_to_sqlite(lst_xl=)
    # clear_base(table='')