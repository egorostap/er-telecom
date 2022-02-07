# необходимо установить pip install pandas, pip install openpyxl, pip install gspread, pip install gspread_dataframe==3.2.2
# эксель таблицы загружены в бд sqlite, данная база данных также находится в репозитории проекта
# готовые запросы по домашнему заданию находятся в переменных: "zd_1_1 - zd_1_3"

# import os
# from pandas import DataFrame, to_datetime
# import datetime as dt
import sqlite3
import openpyxl
import pandas as pd
import gspread
from gspread_dataframe import set_with_dataframe


def export_to_sqlite(xl='data_.xlsx', lst_xl='', column=''):
    '''Экспорт данных из xlsx в sqlite'''

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    # prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 's.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    # connect = sqlite3.connect(prj_dir + '/' + base_name)
    connect = sqlite3.connect(base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute(f'CREATE TABLE IF NOT EXISTS {lst_xl} {column}')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook(xl, data_only=True)
    sheet = file_to_read[lst_xl]

    # Цикл по строкам начиная со второй (в первой заголовки)
    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам начиная 1
        for col in range(1, len(column) + 1):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)
        print(data)

        # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы, необходимо менять количество записей вручную соответственно количеству столбцов
        cursor.execute(f"INSERT INTO {lst_xl} VALUES (?, ?, ?, ?, ?, ?, ?);",
                       (data[0], data[1].date(), data[2], data[3], data[4], data[5], data[6]))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base(table):
    '''Очистка базы sqlite'''

    # Имя базы
    base_name = 's.sqlite3'

    connect = sqlite3.connect(base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute(f"DELETE FROM {table}")
    connect.commit()
    connect.close()


def input_base(sql_request):
    '''SQL запрос возвращает таблицу'''

    # Имя базы
    base_name = 's.sqlite3'
    # соединение с базой
    connect = sqlite3.connect(base_name)

    # Запрос в базу, сохранение и закрытие соединения
    df = pd.read_sql(sql_request, connect)
    # запись в отдельную таблицу бд без зависимостей
    # df_anythink.to_sql(“name_table”, connect)
    connect.commit()
    connect.close()
    return df


# формирование табицы из отдельных sql запросов
# sql_requests_list = []
# headers_list = []
# def make_table(headers_list='', sql_requests_list=''):
#     insights_list = []
#
#     for elem in sql_requests_list:
#         insights_list.append(int(elem))
#     print('данные получены')
#
#     # запись данных в таблицу
#     df = DataFrame()
#     for i in range(len(headers_list)):
#         df[headers_list[i]] = insights_list[i]
#     print('сформирована таблица')
#     return df

def google_sheet_write(df):
    # ACCES GOOGLE SHEET
    sa = gspread.service_account(filename="service_account.json")
    sh = sa.open_by_key('1YDUfbpkDdT6ABQrmBjH79LNHG5pevZPYfs4dwAL2hyo')
    worksheet = sh.worksheet("values")
    print('законектились в гугл шитс')

    # CLEAR SHEET CONTENT
    range_of_cells = worksheet.range('A2:H10000')  # -> Select the range you want to clear
    for cell in range_of_cells:
        cell.value = ''
    worksheet.update_cells(range_of_cells)
    print('таблица чиста')

    # APPEND DATA TO SHEET
    set_with_dataframe(worksheet, df)  # -> THIS EXPORTS YOUR DATAFRAME TO THE GOOGLE SHEET

    print('данные отправлены в гугл таблицы')


'''переменные для загрузки таблиц в бд'''
# переменные с наименованиями вкладок ексель соответственные названия страниц
s_1, s_2, s_3 = 's_1', 's_2', 's_3'
p_1, p_2, p_3, p_4 = 'p_1', 'p_2', 'p_3', 'p_4'

# переменные с наименованиями столбцов в виде картежей по второму заданию, таблицы по первому заданию заводил немного иначе
p_1_column = ('camp_id', 'camp_date', 'channel', 'group_type', 'client_id', 'delivery', 'open')
p_2_column = ('client_id', 'debt_date', 'debt_sum')
p_3_column = ('client_id', 'payment_date', 'payment_sum')
p_4_column = ('client_id', 'app_date', 'app_reason')

# Запуск функции
if __name__ == '__main__':
    pass
    # input_base(sql_request=zd_1_3)
    # export_to_sqlite(lst_xl=p_1, column=p_1_column)
    # clear_base(table=p_1)
